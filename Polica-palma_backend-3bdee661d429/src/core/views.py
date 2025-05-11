import io

import pandas as pd
from barcode import EAN13
from barcode.writer import ImageWriter
from django.http import HttpResponse, HttpResponseNotFound

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from src.factory.enums import ProductFactoryStatus
from src.factory.models import ProductFactory
from src.order.models import Order
from src.product.models import Product


def print_barcode(request, *args, **kwargs):
    product_id = kwargs.get('product_id')
    product = get_object_or_404(Product, pk=product_id)

    if not product.is_barcode_printable():
        return HttpResponseNotFound()

    code = product.code[:-1]

    writer = ImageWriter(format="JPEG")
    options = {
        'module_height': 5,
    }

    EAN13(code, writer=writer).save(f"media/barcode/{product.code}", options=options)
    # writer.set_options(
    #     {
    #         'module_height': 0,
    #         'foreground': 'red'
    #     }
    # )
    # writer.foreground = 'red'
    # writer.module_height = 1
    # writer.module_width = 20
    # with open(f"media/barcode/{product.code}.jpeg", "wb") as f:

    # with open(f"static_files/assets/barcode/{product.code}.jpeg", "rb") as f:
    #     response = HttpResponse(f.read(), content_type="image/jpeg")

    # Delete the image file after serving it to the user
    # os.remove(f"static_files/assets/barcode/{product.code}.jpeg")

    return render(request, 'barcode.html', context={
        "file_path": f"/media/barcode/{product.code}.jpeg",
        "product": product,
        "code": code
    })


def print_receipt(request, *args, **kwargs):
    order_id = kwargs.get('order_id')
    order: Order = get_object_or_404(Order, pk=order_id)
    order_items = order.order_items.all().select_related('product')
    order_item_product_factories = order.order_item_product_factory_set.filter(is_returned=False)

    context = {
        'order': order,
        'order_items': order_items,
        'order_item_product_factories': order_item_product_factories,
        'order_item_count': order_items.count()
    }
    return render(request, 'order_receipt.html', context)


def print_factory_receipt(request, *args, **kwargs):
    factory_id = kwargs.get('factory_id')
    product_factory: ProductFactory = get_object_or_404(ProductFactory, pk=factory_id)

    if product_factory.status not in [
        ProductFactoryStatus.SOLD,
        ProductFactoryStatus.FINISHED,
        ProductFactoryStatus.WRITTEN_OFF
    ]:
        return HttpResponseNotFound()

    if not product_factory.product_code or len(product_factory.product_code) != 13:
        return HttpResponseNotFound()

    code = product_factory.product_code[:-1]

    writer = ImageWriter(format="JPEG")
    options = {
        'module_height': 5,
    }

    EAN13(code, writer=writer).save(f"media/barcode/factories/{product_factory.product_code}", options=options)

    factory_items = product_factory.product_factory_item_set.all().select_related('warehouse_product__product')

    context = {
        'product_factory': product_factory,
        'factory_items': factory_items,
        'barcode_path': f"/media/barcode/factories/{product_factory.product_code}.jpeg",
    }
    return render(request, 'flower_receipt.html', context)


def perfume_products_print(request, *args, **kwargs):
    industry_id = kwargs.get('industry_id')
    products = Product.objects.filter(category__industry_id=industry_id)
    column_names = ['Название', 'Цена']
    columns_data = [(product.name, f"{product.price} Сум") for product in products]
    df = pd.DataFrame(columns=column_names, data=columns_data)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    return response


def industry_products_print(request, *args, **kwargs):
    industry_id = kwargs.get('industry_id')
    products = Product.objects.filter(
        category__industry_id=industry_id,
        is_deleted=False,
    ).with_in_stock().exclude(in_stock=0)
    column_names = ['Название', 'Цена', 'Код', 'Текущее кол-во', 'Фактическое кол-во']
    columns_data = [(product.name, f"{product.price} Сум", product.code, product.in_stock, '') for
                    product in products]
    df = pd.DataFrame(columns=column_names, data=columns_data)

    thin_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        for col_idx, col in enumerate(df.columns, start=1):
            max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2  # +2 for padding
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = max_length

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    return response


def product_factories_export(request, *args, **kwargs):
    category_filter = request.GET.getlist('category', None)

    products = ProductFactory.objects.get_available().select_related(
        'category', 'florist'
    ).filter(status=ProductFactoryStatus.FINISHED).order_by('-created_at')
    if category_filter and not category_filter == []:
        products = products.filter(category__in=category_filter)

    column_names = [
        '#', 'Название', 'Тип Продажи', 'Флорист',
    ]
    columns_data = [
        (
            i + 1,
            p.name,
            p.get_sales_type_display(),
            p.florist.get_full_name(),
        ) for i, p in enumerate(products)
    ]

    df = pd.DataFrame(columns=column_names, data=columns_data)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        start_row = 3
        end_row = start_row + len(df)

        df.to_excel(writer, index=False, sheet_name='Buketi', startrow=start_row - 1)
        ws = writer.sheets['Buketi']

        cell_format_props = {
            'border': 2
        }
        cell_head_format_props = {
            'border': 2,
            'bg_color': '#ddebf7',
            'bold': True,
        }

        cell_format_e = writer.book.add_format(
            {
                "align": "center",
                "font_size": 14,
                'bold': True
            }
        )
        ws.merge_range("A1:G1",
                       f"Отчет по существующим букетам от "
                       f"{timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')}",
                       cell_format_e)
        cell_format = writer.book.add_format(
            cell_format_props
        )
        cell_format_head = writer.book.add_format(
            cell_head_format_props
        )
        ws.conditional_format(f'A{start_row}:D{end_row}', {
            'type': 'cell',
            'criteria': '>=',
            'value': 0,
            'format': cell_format,
        })
        ws.conditional_format(f'A{start_row}:D{start_row}', {
            'type': 'cell',
            'criteria': '>=',
            'value': 0,
            'format': cell_format_head,
        })
        ws.set_column(0, 0, 5)
        ws.set_column(1, 2, 20)
        ws.set_column(3, 3, 40)

    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=buket.xlsx'
    return response


class ProductFactoriesPrintView(TemplateView):
    template_name = 'product_factories_print.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['products'] = ProductFactory.objects.get_available().select_related(
            'category', 'florist'
        ).filter(status=ProductFactoryStatus.FINISHED).order_by('-created_at')
        return context
